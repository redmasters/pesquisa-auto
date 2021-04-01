from openpyxl import Workbook
from drive import Driver

URL = "https://lista.mercadolivre.com.br/nintendo-switch#D[A:nintendo%20switch]"
navegador = Driver() 

navegador.go_url(URL)

file = open("links.txt", "w+")

for link in navegador.get_links():
    file.write(link + "\n")
file.close

file = open("links.txt", "r+")

wb = Workbook()
ws = wb.active

ws['A1'] = 'Title'
ws['B1'] = 'Price'
ws['C1'] = 'URL'

pos = 2

for line in file:
    if len(line) > 20:
        data = navegador.get_data(line.strip())
        ws['A{}'.format(pos)] = data['item_title']
        ws['B{}'.format(pos)] = data['item_price']
        ws['C{}'.format(pos)] = line
    pos += 1
    
wb.save('saida.xlsm')

navegador.fechar()

# page = requests.get(URL)

# soup = BeautifulSoup(page.content, 'html5lib')
# # print(soup.prettify())

# title = soup.find('a', class_='ui-search-link')

# print(title)
